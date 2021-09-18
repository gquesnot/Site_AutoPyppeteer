import json
import os
from time import time

from openpyxl import load_workbook, Workbook
import string
from django.templatetags.static import static

class Excel:
    alphabet = list(string.ascii_uppercase)

    def __init__(self, path=None):
        self.type = "out" if path is None else "in"
        self.path = path
        if self.type == "in":
            self.book = load_workbook(filename=f"{os.getcwd()}/{path}")
            self.sheet = self.book[self.getSheetNames()[0]]
            self.test = lambda start,end,sheetName=None: [{self.alphabet[i]: cell.value for i, cell in enumerate(row)} for row in (self.book[sheetName] if sheetName is not None else self.sheet).iter_rows(min_row=int(start.split(':')[1]), max_row=int(end.split(':')[1]), min_col=self.alphabet.index(start.split(':')[0]), max_col=self.alphabet.index(end.split(':')[0]))]
        else:
            self.book = Workbook()
            self.sheet = self.book['Sheet']

    def getSheetNames(self):
        return self.book.get_sheet_names()

    def selectSheet(self, sheetName):
        self.sheet = self.book[sheetName]

    def insertValues(self, datas):

        lenDataKey = len(datas["0"].keys())
        for i ,data in datas.items():

            for k,v in data.items():
                self.sheet[f"{k}{int(i) + 1}"] = v
        rPath= f"{os.getcwd()}/static/filesOut/{str(int(time()))}.xlsx"

        self.book.save(rPath)
        path = rPath.split('/')
        path = os.path.join(path[-2], path[-1])
        url = static(path)

        return url

    def getValues(self, start, end, sheetName=None):
        try:
            result = [{self.alphabet[i]: cell.value for i, cell in enumerate(row)} for row in (self.book[sheetName] if sheetName is not None else self.sheet).iter_rows(min_row=int(start.split(':')[1]), max_row=int(end.split(':')[1]), min_col=self.alphabet.index(start.split(':')[0]), max_col=self.alphabet.index(end.split(':')[0]) + 1)]
            # sheet = self.book[sheetName] if sheetName is not None else self.sheet
            # startSplited = start.split(':')
            # endSplited = end.split(':')
            # minCol, minRow = self.alphabet.index(startSplited[0]), int(startSplited[1])
            # maxCol , maxRow = self.alphabet.index(endSplited[0]), int(endSplited[1])
            # result = []
            # for row in self.sheet.iter_rows(min_row=min, max_row=maxRow, min_col=minCol, max_col=maxCol):
            #     tmp = dict()
            #     for i, cell in enumerate(row):
            #         tmp[self.alphabet[i]] = cell.value
            #     result.append(tmp)
            return result
        except:
            return None


if __name__ == '__main__':
    ex = Excel('avancement_intégration_site_XENOM_SOLUTIONS.xlsx')
    ex.selectSheet('contenus produit')
    print(json.dumps(ex.test("A:1", "E:7"), indent=4))
